#!/usr/bin/env python3
"""Build a spreadsheet view from state/. The CSVs are the source of truth; this
is a read-only rendering for humans.

Usage:
    python scripts/render_tracker.py [--out tracker.xlsx]

Edits made in the spreadsheet do not flow back. One writer, one truth: a fast
agent and a human both editing the same file is how a tracker gets corrupted.

Tabs:
    Prospects  - one row per company, with derived current status
    People     - contacts, channels, warmth
    Funnel     - counts derived from the event log, not stored anywhere
    Events     - the raw log, most recent first
"""
import argparse
import csv
import datetime as _dt
import pathlib
import sys
from collections import defaultdict

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from common import ROOT, read_events, state_path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: openpyxl. Run: pip install -r requirements.txt")

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# The CSVs are the machine's format: plain, diffable, append-safe, cheap for an
# agent to read. This spreadsheet is the human's format, so it gets the colour.
# Same data, two audiences.
STATUS_STYLE = {
    "closed_won":     ("C6F6D5", "22543D", True),
    "meeting_booked": ("C6F6D5", "22543D", True),
    "reply_received": ("BEE3F8", "1A365D", True),
    "reply_drafted":  ("BEE3F8", "1A365D", False),
    "follow_up_sent": ("E9D8FD", "44337A", False),
    "sent":           ("EDF2F7", "2D3748", False),
    "bounced":        ("FED7D7", "742A2A", False),
    "verify_failed":  ("FED7D7", "742A2A", False),
    "not_interested": ("FEEBC8", "7B341E", False),
    "closed_lost":    ("FEEBC8", "7B341E", False),
    "do_not_contact": ("1A202C", "FFFFFF", True),
    "excluded":       ("1A202C", "FFFFFF", True),
    "needs_named_contact": ("FEFCBF", "744210", False),
    "skipped":        ("FEFCBF", "744210", False),
    "not contacted":  ("FFFFFF", "718096", False),
}

TIER_FILL = {"1": "C6F6D5", "2": "E9D8FD", "3": "EDF2F7", "4": "F7FAFC"}


def style_status_cell(cell, status):
    fill, font, bold = STATUS_STYLE.get(status, ("FFFFFF", "2D3748", False))
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=font, bold=bold)


def read_csv(name):
    path = state_path(name)
    if not path.exists():
        return []
    with open(path, newline="") as fh:
        return list(csv.DictReader(fh))


def current_status(events_by_company):
    """Current state is the most recent event, derived rather than stored, so it
    cannot drift out of sync with what actually happened."""
    return {company: sorted(evs, key=lambda e: e["timestamp"])[-1]
            for company, evs in events_by_company.items()}


def funnel(events):
    by_company = defaultdict(set)
    for e in events:
        by_company[e["company"]].add(e["action"])

    total = len(by_company)
    sent = {c for c, a in by_company.items() if a & {"sent", "follow_up_sent"}}
    bounced = {c for c, a in by_company.items() if "bounced" in a}
    delivered = sent - bounced
    followed = {c for c, a in by_company.items() if "follow_up_sent" in a}
    replied = {c for c, a in by_company.items() if "reply_received" in a}
    negative = {c for c, a in by_company.items() if a & {"not_interested", "closed_lost"}}
    meetings = {c for c, a in by_company.items() if "meeting_booked" in a}
    skipped = {c for c, a in by_company.items() if "skipped" in a}

    def pct(n, d):
        return f"{n / d * 100:.1f}%" if d else "n/a"

    return [
        ("Companies in the log", total, ""),
        ("Attempted", len(sent), ""),
        ("Bounced", len(bounced), pct(len(bounced), len(sent))),
        ("Delivered", len(delivered), pct(len(delivered), len(sent))),
        ("Got at least one follow-up", len(followed), pct(len(followed), len(delivered))),
        ("Any human response", len(replied), pct(len(replied), len(delivered))),
        ("Explicit no", len(negative), pct(len(negative), len(delivered))),
        ("Meetings booked", len(meetings), pct(len(meetings), len(delivered))),
        ("Skipped (not sent, with a reason)", len(skipped), ""),
    ]


def write_sheet(ws, headers, rows, widths=None, status_col=None, tier_col=None):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    if status_col:
        idx = headers.index(status_col) + 1
        for r in range(2, ws.max_row + 1):
            style_status_cell(ws.cell(row=r, column=idx),
                              str(ws.cell(row=r, column=idx).value or ""))
    if tier_col and tier_col in headers:
        idx = headers.index(tier_col) + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            shade = TIER_FILL.get(str(cell.value or "").strip())
            if shade:
                cell.fill = PatternFill("solid", fgColor=shade)
                cell.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for i, header in enumerate(headers, start=1):
        width = (widths or {}).get(header, min(max(len(str(header)) + 4, 14), 46))
        ws.column_dimensions[get_column_letter(i)].width = width


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(ROOT / "tracker.xlsx"))
    args = ap.parse_args()

    prospects = read_csv("prospects.csv")
    people = read_csv("people.csv")
    events = read_events()

    by_company = defaultdict(list)
    for e in events:
        by_company[e["company"]].append(e)
    latest = current_status(by_company)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Prospects"
    write_sheet(
        ws,
        ["Company", "Cohort", "Fit", "Country", "Score", "Tier",
         "Current status", "Last action", "Last touch", "Scoring reasoning"],
        [[p.get("company"), p.get("group"), p.get("product_fit"), p.get("country"),
          p.get("score"), p.get("tier"),
          latest.get(p.get("company"), {}).get("action", "not contacted"),
          latest.get(p.get("company"), {}).get("detail", ""),
          latest.get(p.get("company"), {}).get("timestamp", "")[:10],
          p.get("scoring_reasoning")]
         for p in prospects],
        widths={"Scoring reasoning": 70, "Last action": 40},
        status_col="Current status", tier_col="Tier",
    )

    write_sheet(
        wb.create_sheet("People"),
        ["Company", "Name", "Title", "Role type", "Email", "Verdict", "Channel",
         "Mutual connections", "Warmth", "Notes"],
        [[p.get("company"), p.get("name"), p.get("title"), p.get("role_type"),
          p.get("email"), p.get("email_verdict"),
          p.get("other_channel") or ("linkedin" if p.get("linkedin") else "email"),
          p.get("mutual_connections"), p.get("warmth"), p.get("notes")]
         for p in people],
    )

    fs = wb.create_sheet("Funnel")
    write_sheet(fs, ["Metric", "Count", "Rate"], [list(r) for r in funnel(events)],
                widths={"Metric": 38})
    start = fs.max_row + 3
    fs.cell(row=start, column=1, value="Status colours").font = Font(bold=True)
    for i, (status, _) in enumerate(sorted(STATUS_STYLE.items()), start=start + 1):
        cell = fs.cell(row=i, column=1, value=status)
        style_status_cell(cell, status)

    write_sheet(
        wb.create_sheet("Events"),
        ["Timestamp", "Company", "Action", "Channel", "Email", "Detail", "Run"],
        [[e.get("timestamp"), e.get("company"), e.get("action"), e.get("channel"),
          e.get("person_email"), e.get("detail"), e.get("run_id")]
         for e in sorted(events, key=lambda x: x["timestamp"], reverse=True)],
        widths={"Detail": 60}, status_col="Action",
    )

    wb.save(args.out)
    print(f"wrote {args.out}")
    print(f"  {len(prospects)} prospects, {len(people)} people, {len(events)} events")
    print("  This file is a view. Edit state/*.csv, not the spreadsheet.")


if __name__ == "__main__":
    main()
